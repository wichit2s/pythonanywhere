from landmark.models import *
from openpyxl import Workbook, load_workbook

def save():
    wb = Workbook()

    categoryws = wb.create_sheet(title='Category')
    categoryws.append(['id', 'name'])
    for x in Category.objects.all():
        categoryws.append([ x.id, x.name ])

    landmarkws = wb.create_sheet(title='Landmark')
    landmarkws.append(['id', 'image', 'title', 'description', 'rating', 'reviews'])
    for x in Landmark.objects.all():
        landmarkws.append(
            [ x.id, x.image, x.title, x.description, x.rating, x.reviews ]
            )
    wb.save('landmarks.xlsx')

def load():
    wb = load_workbook('landmarks.xlsx')
    ws = wb['Category']
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=2, values_only=True):
        if row[0] == None:
            break
        else:
            c = Category(id=row[0], name=row[1])
            c.save()

    ws = wb['Landmark']
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=2, values_only=True):
        if row[0] == None:
            break
        else:
            c = Category(id=row[0], name=row[1])
            c.save()

